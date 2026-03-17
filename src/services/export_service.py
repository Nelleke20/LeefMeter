"""Excel export service for LeefMeter activity data.

Uses the Repository pattern to access activity data and writes two sheets:
- "Ingevulde dagen": time-grid with days as columns and 30-min rows.
- "Per categorie": unique activity names grouped by category.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models.activity import Activity
from src.models.template import Template
from src.repositories.base import ActivityRepository
from src.services.template_service import TemplateService
from src.storage import get_data_dir

_HEADER_FILL = PatternFill(start_color="1A7F6F", end_color="1A7F6F", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FONT = Font(bold=True)


def get_export_path() -> Path:
    """Return the platform-appropriate path for the Excel export file.

    On Android, tries to save to /storage/emulated/0/Documents/ (the public
    Documents folder visible in the Files app). This requires the
    MANAGE_EXTERNAL_STORAGE permission to be granted in Android Settings.

    Falls back to the app's external files directory if Documents is not
    accessible, then to internal app storage as a last resort.

    On desktop saves to ~/Downloads.

    Returns:
        Path to leefmeter_export.xlsx.
    """
    import os

    if os.environ.get("FLET_APP_STORAGE_DATA"):
        # Primary: public Documents folder (requires MANAGE_EXTERNAL_STORAGE)
        docs = Path("/storage/emulated/0/Documents")
        try:
            docs.mkdir(parents=True, exist_ok=True)
            test = docs / ".leefmeter_test"
            test.touch()
            test.unlink()
            return docs / f"{date.today().strftime('%Y%m%d')}_leefmeter_export.xlsx"
        except OSError:
            pass
        # Fallback: app-specific external dir (no permissions needed)
        env = os.environ.get("FLET_APP_STORAGE_DATA", "")
        for part in Path(env).parts:
            if "." in part and not part.startswith("/"):
                external = Path(f"/sdcard/Android/data/{part}/files")
                try:
                    external.mkdir(parents=True, exist_ok=True)
                    stamp = date.today().strftime("%Y%m%d")
                    return external / f"{stamp}_leefmeter_export.xlsx"
                except OSError:
                    break
        stamp = date.today().strftime("%Y%m%d")
        return get_data_dir() / f"{stamp}_leefmeter_export.xlsx"
    stamp = date.today().strftime("%Y%m%d")
    downloads = Path.home() / "Downloads"
    downloads.mkdir(parents=True, exist_ok=True)
    return downloads / f"{stamp}_leefmeter_export.xlsx"


_CAT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_CAT_FONT = Font(bold=True, color="000000")

_DUTCH_WEEKDAYS: tuple[str, ...] = ("ma", "di", "wo", "do", "vr", "za", "zo")

# Time-grid constants
_SLOT_START_HOUR: int = 6
_SLOT_END_HOUR: int = 22
_SLOT_MINUTES: int = 30


def _day_header(d: date) -> str:
    """Format a date as a Dutch weekday-abbreviated column header.

    Args:
        d: The date to format.

    Returns:
        String like "ma 15-03".
    """
    return f"{_DUTCH_WEEKDAYS[d.weekday()]} {d.strftime('%d-%m')}"


def _all_slots() -> list[str]:
    """Return all 30-min time slot labels from 06:00 to 22:00 inclusive.

    Returns:
        List of "HH:MM" strings, e.g. ["06:00", "06:30", ..., "22:00"].
    """
    slots: list[str] = []
    idx = 0
    while True:
        total = _SLOT_START_HOUR * 60 + idx * _SLOT_MINUTES
        h, m = divmod(total, 60)
        label = f"{h:02d}:{m:02d}"
        slots.append(label)
        if h == _SLOT_END_HOUR and m == 0:
            break
        idx += 1
    return slots


def _slots_for_activity(activity: Activity) -> list[str]:
    """Return all 30-min slot labels occupied by an activity.

    Args:
        activity: The activity to calculate slots for.

    Returns:
        List of "HH:MM" slot labels, or empty list if start_time is not set.
    """
    if not activity.start_time:
        return []
    try:
        start_dt = datetime.strptime(activity.start_time, "%H:%M")
    except ValueError:
        return []
    occupied: list[str] = []
    current = start_dt
    end_dt = start_dt + timedelta(minutes=activity.duration_minutes)
    while current < end_dt:
        occupied.append(current.strftime("%H:%M"))
        current += timedelta(minutes=_SLOT_MINUTES)
    return occupied


class ExportService:
    """Creates Excel exports of LeefMeter activity data.

    Produces two worksheets:
    - "Ingevulde dagen": all logged days listed sequentially.
    - "Per categorie": unique activity names grouped by category.
    """

    def __init__(
        self, repository: ActivityRepository, template_service: TemplateService
    ) -> None:
        """Initialise with an activity repository and template service.

        Args:
            repository: Source of activity data to export.
            template_service: Source of registered activity templates.
        """
        self._repository = repository
        self._template_service = template_service

    def export(
        self,
        output_path: Path | None = None,
        from_date: date | None = None,
        to_date: date | None = None,
    ) -> Path:
        """Export activities to an Excel file with two sheets.

        Sheet 1 "Weekschema": time-grid with one column per day.
        Sheet 2 "Per categorie": activities grouped by category.

        Args:
            output_path: Where to save the Excel file. Defaults to a
                platform-appropriate path (app storage on Android, ~/Downloads
                on desktop).
            from_date: Include activities on or after this date.
            to_date: Include activities on or before this date.

        Returns:
            Path to the saved Excel file.
        """
        resolved_path = output_path if output_path is not None else get_export_path()
        activities = sorted(
            self._repository.get_all(), key=lambda a: (a.date, a.start_time or "")
        )
        if from_date:
            activities = [a for a in activities if a.date >= from_date]
        if to_date:
            activities = [a for a in activities if a.date <= to_date]

        wb = Workbook()
        days_ws: Worksheet = wb.active  # type: ignore[assignment]
        days_ws.title = "Ingevulde dagen"
        self._write_ingevulde_dagen_sheet(days_ws, activities)

        cat_ws: Worksheet = wb.create_sheet(title="Per categorie")
        templates = self._template_service.get_all_templates()
        self._write_per_categorie_sheet(cat_ws, templates)

        resolved_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(resolved_path)
        return resolved_path

    @staticmethod
    def _write_ingevulde_dagen_sheet(
        ws: Worksheet, activities: Sequence[Activity]
    ) -> None:
        """Write the Ingevulde-dagen sheet as a time-grid.

        Row 1 has "Tijd" in column A and one column per unique date.
        Rows 2..N have one 30-min slot per row from 06:00 to 22:00.
        The final row shows total points per day.

        Args:
            ws: The openpyxl worksheet to write to.
            activities: Activities to render, pre-sorted by date.
        """
        sorted_dates = sorted({a.date for a in activities})

        # Header row
        tijd_cell = ws.cell(row=1, column=1, value="Tijd")
        tijd_cell.font = _HEADER_FONT
        tijd_cell.fill = _HEADER_FILL
        tijd_cell.alignment = Alignment(horizontal="center")

        for col_idx, d in enumerate(sorted_dates, start=2):
            cell = ws.cell(row=1, column=col_idx, value=_day_header(d))
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Build lookup: (date, slot) → activity name
        slot_map: dict[tuple[date, str], str] = {}
        for act in activities:
            for slot in _slots_for_activity(act):
                key = (act.date, slot)
                if key not in slot_map:
                    slot_map[key] = act.name

        # Slot rows
        slots = _all_slots()
        for row_offset, slot_label in enumerate(slots, start=2):
            ws.cell(row=row_offset, column=1, value=slot_label).alignment = Alignment(
                horizontal="center"
            )
            for col_idx, d in enumerate(sorted_dates, start=2):
                value = slot_map.get((d, slot_label), "")
                if value:
                    ws.cell(row=row_offset, column=col_idx, value=value)

        # Totals row
        total_row = len(slots) + 2
        ws.cell(row=total_row, column=1, value="Totaal").font = _TOTAL_FONT

        points_by_date: dict[date, int] = defaultdict(int)
        for act in activities:
            points_by_date[act.date] += act.points

        for col_idx, d in enumerate(sorted_dates, start=2):
            cell = ws.cell(row=total_row, column=col_idx, value=points_by_date[d])
            cell.font = _TOTAL_FONT
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 8
        for col_idx in range(2, len(sorted_dates) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    @staticmethod
    def _write_per_categorie_sheet(
        ws: Worksheet, templates: Sequence[Template]
    ) -> None:
        """Write the Per-categorie sheet with registered activity templates per category.

        Each section has a category header row followed by the registered activity
        names in that category, sorted alphabetically.

        Args:
            ws: The openpyxl worksheet to write to.
            templates: Registered activity templates to render.
        """
        by_category: dict[str, list[str]] = defaultdict(list)
        for tmpl in templates:
            by_category[tmpl.category].append(tmpl.name)

        category_order = ["rust", "laag", "gemiddeld", "zwaar"]
        current_row = 1
        for category in category_order:
            names = sorted(by_category.get(category, []))
            if not names:
                continue

            # Category header row
            cat_cell = ws.cell(
                row=current_row, column=1, value=f"Categorie: {category}"
            )
            cat_cell.font = _CAT_FONT
            cat_cell.fill = _CAT_FILL
            current_row += 1

            # One row per unique activity name
            for name in names:
                ws.cell(row=current_row, column=1, value=name)
                current_row += 1

            # Blank separator row
            current_row += 1

        # Column width
        ws.column_dimensions["A"].width = 32
